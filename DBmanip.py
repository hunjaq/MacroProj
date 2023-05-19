##backend database manipulation
#utilize database for quick search, sort
#delete database ans save to spreadsheet to save space when not used
#update spreadsheet when inserting/removing
#no repeats, 
import xlrd
import sqlite3 
from openpyxl import Workbook, load_workbook


"""
test method to print database
"""
def test_all(db):
    con = sqlite3.connect("fooddb.db")  # create db
    cur = con.cursor()  # create cursor
    command = "SELECT * FROM {}".format(db)
    res = cur.execute(command)
    print(res.fetchall())
    
    
"""
removes object with given name from db
"""
def remove_from_database(name, db):
    con = sqlite3.connect("fooddb.db")  # create db
    cur = con.cursor()  # create cursor
    
    #test if in database
    #res = cur.execute("SELECT name FROM {} WHERE NOT EXISTS (SELECT 1 FROM {} WHERE name = \'{}\')".format(db, db, name))
    #res = cur.execute("NOT EXISTS (SELECT 1 FROM {} WHERE name = \'{}\'".format(db, name))
    res = cur.execute("SELECT name FROM {} WHERE name = \'{}\'".format(db, name))
    if (res == False):
        print("{} does not exist in {} database".format(name, db))
        return
    command = "DELETE FROM {} WHERE name = \'{}\'".format(db, name)
    cur.execute(command)
    con.commit() #commit changes
    print("{} removed from table {}".format(name, db))
    
    """ **test
    res = cur.execute("SELECT name FROM comps")
    print(res.fetchall())
    """
    con.close() #save db
    
    #update excel
    wb = load_workbook("Macros.xlsx")
    if (db == "comps"): #remove from comps sheet
        sheet = wb['Components']
    else: #remove from foods
        sheet = wb['Meals']
    #find where in sheet
    row = 1
    test_name = sheet['A1'] #start at first val
    while (test_name != name and row <= num_rows):
        row = row + 1
        test_name = sheet['A{}'.format(row)]
    if (row == num_rows+1):
        print("{} does not exist in {} database".format(name, db))
    
    
    """
    add to comps the given object
    """
def add_to_comps(name, cals, carbs, protein, fat, serving):
    #utilize sql for database
    con = sqlite3.connect("fooddb.db")  # create db
    cur = con.cursor()  # create cursor
    
    #insert
    command = "INSERT INTO comps VALUES (\'{}\', {}, {}, {}, {}, {})".format(name, cals, carbs, protein, fat, serving)
    cur.execute(command)
    con.commit()
    print("{} added to comps database".format(name))
    con.close()
    
    #update excel
    wb = load_workbook('Macros.xlsx')
    sheet = wb['Components'] #access components sheet
    #add to bottom, each column
    sheet['A{}'.format(num_rows+1)].value = name
    sheet['B{}'.format(num_rows+1)].value = cals
    sheet['C{}'.format(num_rows+1)].value = carbs
    sheet['D{}'.format(num_rows+1)].value = protein
    sheet['E{}'.format(num_rows+1)].value = fat
    sheet['F{}'.format(num_rows+1)].value = serving
    #update file
    wb.save('Macros.xlsx')
    
    
"""
search db with given condition - cals, name, etc
upper = upper bound, lower = lower bound
"""
def search_range(db, condition, lower, upper):
    print()
    con = sqlite3.connect("fooddb.db")  # create db
    cur = con.cursor()  # create cursor
    command = "SELECT * FROM \'{}\' WHERE {} BETWEEN {} AND {}".format(db, condition, lower, upper)
    #get all details from db with lower and upper constraints
    res = cur.execute(command)
    #print(res.fetchall())
    #con.close()
    return res.fetchall()
    con.close()
    
    
"""
search db with given condition, results ascending
"""
def search_asc(db, condition):
    print()
    con = sqlite3.connect("fooddb.db")  # create db
    cur = con.cursor()  # create cursor
    
    command = "SELECT * FROM \'{}\' ORDER BY {} ASC".format(db, condition)
    res = cur.execute(command)
    return res.fetchall()
    con.close()
    
    
"""
search db with given condition, results descending
"""
def search_desc(db, condition):
    print()
    con = sqlite3.connect("fooddb.db")  # create db
    cur = con.cursor()  # create cursor
    command = "SELECT * FROM \'{}\' ORDER BY {} DESC".format(db, condition)
    res = cur.execute(command)
    return res.fetchall()
    con.close()
    
    
"""
takes from component excel sheet and populates db
"""
def create_comp_database():
    
    # utilize sql for database
    con = sqlite3.connect("fooddb.db")  # create db
    cur = con.cursor()  # create cursor

    # clean
    cur.execute("DROP TABLE comps")
    # create components table, don't run multiple times?
    cur.execute("CREATE TABLE comps(name, cals, carbs, protein, fat, serving_size)")

    # # insert components into database
    row = 1  # start on row 3
    col = 0  # start on first col
    #numrows = worksheet
    for x in range (10):
        name = worksheet.cell_value(row, col)
        # correct spaces
        name = name.replace(" ", "_")      
        cals = worksheet.cell_value(row, col + 1)
        carbs = worksheet.cell_value(row, col + 2)
        protein = worksheet.cell_value(row, col + 3)
        fat = worksheet.cell_value(row, col + 4)
        serving = worksheet.cell_value(row, col + 5)
        # price = worksheet.cell_value(row, col+6)
        command = "INSERT INTO comps VALUES (\'{}\', {}, {}, {}, {}, {})".format(name, cals, carbs, protein, fat, serving)
        #print(command)
        cur.execute(command) #execute insert
        row = row + 1 #iterate to next row
    con.commit() #commit additions
    res = cur.execute("SELECT cals FROM comps")
    print(res.fetchall())
    cur.close() #write to disk



# create foods table
# cur.execute("CREATE TABLE foods(name, components)")


# open workbook
workbook = xlrd.open_workbook("Macros.xlsx")
# open component worksheet
worksheet = workbook.sheet_by_index(0)  # has one sheet
num_rows = worksheet.nrows
#print(num_rows)
wb = load_workbook('Macros.xlsx')

create_comp_database() #populate components
remove_from_database("pasta", "comps")
remove_from_database("not_there", "comps")
#print(search_range('comps', 'cals', 200, 300))
#print(search_asc('comps', 'protein'))
#add_to_comps("test", 1, 2, 3, 4, 5)
#test_all("comps")
